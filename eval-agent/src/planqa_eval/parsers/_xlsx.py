from __future__ import annotations

from collections.abc import Iterator
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet


def find_header_row(
    ws: Worksheet, required_column: str, max_scan_rows: int = 5
) -> tuple[int, dict[str, int]]:
    """Scan the first few rows for the one containing `required_column`, rather than
    assuming headers always sit on row 1 — Review1-6 sheets have two banner rows above
    their real header row."""
    for row in range(1, max_scan_rows + 1):
        headers = {
            ws.cell(row=row, column=col).value: col
            for col in range(1, ws.max_column + 1)
            if ws.cell(row=row, column=col).value
        }
        if required_column in headers:
            return row, headers
    raise ValueError(f"no header row containing {required_column!r} in sheet {ws.title!r}")


def iter_rows(ws: Worksheet, header_row: int, headers: dict[str, int]) -> Iterator[dict[str, Any]]:
    for row in range(header_row + 1, ws.max_row + 1):
        values = {name: ws.cell(row=row, column=col).value for name, col in headers.items()}
        if any(v is not None for v in values.values()):
            yield values


def cell_str(values: dict[str, Any], key: str) -> str | None:
    value = values.get(key)
    if value is None:
        return None
    text = str(value).strip()
    return text or None
