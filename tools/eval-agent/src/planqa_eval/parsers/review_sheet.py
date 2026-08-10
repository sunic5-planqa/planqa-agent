from __future__ import annotations

from pathlib import Path

import openpyxl

from planqa_eval.parsers._xlsx import cell_str, find_header_row, iter_rows
from planqa_schemas.schema import Issue


def parse_review_sheets(xlsx_path: Path) -> dict[str, list[Issue]]:
    """Parses every sheet whose name starts with "Review" (Review1, Review2, ...) rather
    than a fixed count, so an empty Review5/6 or a future Review7 need no code change.
    These are the human-reviewer baselines used by the Aggregator, not the review agent."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    by_sheet: dict[str, list[Issue]] = {}
    for name in wb.sheetnames:
        if not name.startswith("Review"):
            continue
        ws = wb[name]
        try:
            header_row, headers = find_header_row(ws, "Document ID")
        except ValueError:
            by_sheet[name] = []
            continue

        issues: list[Issue] = []
        for values in iter_rows(ws, header_row, headers):
            doc_id = cell_str(values, "Document ID")
            if not doc_id:
                continue
            original_text = cell_str(values, "원문")
            rationale = cell_str(values, "근거")
            issues.append(
                Issue(
                    doc_id=doc_id,
                    level=cell_str(values, "Level") or "",
                    rule_id=cell_str(values, "Rule ID") or "",
                    location=cell_str(values, "위치") or "",
                    description=rationale or original_text or "",
                    source=f"review_sheet:{name}",
                    issue_id=cell_str(values, "Issue ID"),
                    original_text=original_text,
                    rationale=rationale,
                    fix_direction=cell_str(values, "수정 방향"),
                )
            )
        by_sheet[name] = issues
    return by_sheet
