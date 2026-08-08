from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

from planqa_eval.parsers._xlsx import cell_str, find_header_row, iter_rows

SHEET_NAME = "Documents"


@dataclass(frozen=True, slots=True)
class DocumentMeta:
    doc_id: str
    project_name: str
    domain: str
    title: str
    version: str


def parse_documents(xlsx_path: Path) -> dict[str, DocumentMeta]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]
    header_row, headers = find_header_row(ws, "Document ID")

    documents: dict[str, DocumentMeta] = {}
    for values in iter_rows(ws, header_row, headers):
        doc_id = cell_str(values, "Document ID")
        if not doc_id:
            continue
        documents[doc_id] = DocumentMeta(
            doc_id=doc_id,
            project_name=cell_str(values, "프로젝트명") or "",
            domain=cell_str(values, "서비스 분야") or "",
            title=cell_str(values, "문서명") or "",
            version=cell_str(values, "버전") or "",
        )
    return documents


def load_source_text(doc_id: str, source_dir: Path) -> str | None:
    matches = sorted(source_dir.glob(f"{doc_id}_*.md"))
    if not matches:
        return None
    return matches[0].read_text(encoding="utf-8")
