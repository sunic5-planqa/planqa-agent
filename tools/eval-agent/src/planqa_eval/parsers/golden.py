from __future__ import annotations

from pathlib import Path

import openpyxl

from planqa_eval.parsers._xlsx import cell_str, find_header_row, iter_rows
from planqa_schemas.schema import Issue

SHEET_NAME = "golden dataset"


def parse_golden_dataset(xlsx_path: Path) -> list[Issue]:
    """Re-reads the `golden dataset` sheet fresh every call. The sheet is a moving target
    (more rows get confirmed over time) — never cache or hardcode its contents."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]
    header_row, headers = find_header_row(ws, "Document ID")

    issues: list[Issue] = []
    for values in iter_rows(ws, header_row, headers):
        doc_id = cell_str(values, "Document ID")
        if not doc_id:
            continue
        original_text = cell_str(values, "원문")
        rationale = cell_str(values, "근거")
        issues.append(
            Issue(
                doc_id=doc_id,
                level=cell_str(values, "Level") or "",
                rule_id=cell_str(values, "Rule ID") or "",
                location=cell_str(values, "위치") or "",
                description=rationale or original_text or "",
                exception_ref=None,
                source="golden",
                original_text=original_text,
                rationale=rationale,
                fix_direction=cell_str(values, "수정 방향"),
            )
        )
    return issues
